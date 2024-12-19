import numpy as np
import pandas as pd
from openpyxl import Workbook, worksheet
from openpyxl.styles import PatternFill, numbers, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from .sales_order_code_generator import generate_sales_order_str
from .sales_persons import sales_persons_dict
import sys
from datetime import datetime

def generate_sales_import(data: pd.DataFrame, starting_num: int, output_dir  = "../../temp"):
    # create new excel file
    wb = Workbook()
    ws = wb.active

    previous_r = None
    current_si_no = starting_num
    # iterate thru every row of dataframe
    for r in dataframe_to_rows(data, index=False, header=True):
        if r[0] == "SalesOrderCode":
            for j in range(5):
                ws.append(r)
            for row in ws['A2':'BJ6']:
                for cell in row:
                    cell.value = None
            yellow_fill = PatternFill(start_color='FFFF00', fill_type="solid")
            # set to yellow
            for col in ws['A':'BJ']:
                col[4].fill = yellow_fill
            continue
        if previous_r == None:
            r[0] = generate_sales_order_str(current_si_no)
            current_si_no += 1
            previous_r = r
        elif r[9] != previous_r[9]:
            r[0] = generate_sales_order_str(current_si_no)
            current_si_no += 1
            previous_r = r
        else:
            # check stock code first
            for i in range(0,16):
                r[i] = ""
        ws.append(r)

    orange_fill = PatternFill(start_color='D2691E', fill_type="solid")
    for col in ws['A':'AC']:
        col[0].fill = orange_fill

    grey_fill = PatternFill(start_color='808080', fill_type="solid")
    for i in range(1, data.shape[0] + 7):
        ws['AD' + str(i)].fill = grey_fill

    blue_fill = PatternFill(start_color='00BFFF', fill_type="solid")
    for col in ws['AE':'BJ']:
        col[0].fill = blue_fill

    ws['AD1'] = ""

    for row in range(7, data.shape[1]+7):
        if (ws["B" + str(row)].value != None and ws["B" + str(row)].value != ''):
            # format date to `mm/dd/yyyy`
            dt_temp = datetime.strptime(ws["B" + str(row)].value, '%d-%m-%Y %H:%M').strftime('%m/%d/%Y')
            ws["B" + str(row)].value = dt_temp
            ws["D" + str(row)].value = dt_temp
        ws["{}{}".format("J", row)].number_format = numbers.FORMAT_NUMBER

    wb.save(output_dir)
    return True
